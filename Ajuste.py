from lib2to3.pytree import convert
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import openpyxl

def ajustar(inicio):

    print("\n#####################################################\n")
    matricula = float(input("Digite seu número de matrícula:  "))

    #Aqui vamos verificar se o número de matricula digitado existe

    path = "dados_dos_alunos.xlsx"
 
    book = openpyxl.load_workbook(path)

    pagina = book['Página1']

    matriculas = []

    for rows in pagina.iter_rows(min_row=2):
        matriculas.append(rows[2].value)

    i = 0

    verificando_matricula_existe = False

    while (i < len(matriculas)):


        if (matricula == matriculas[i]):

            verificando_matricula_existe = True
            break

        i = i + 1
            

    if verificando_matricula_existe == False:

        print("Seu número de matrícula é inválido!")
        tente_novamente('iniciar')



    print("\n")

    print("As opçõe disponíveis são: \n")
    print("1 - Remover disciplina")
    print("2 - Adicioinar disciplina\n")

    opcao = input("Digite o número correspondente a sua opção: ")

    if opcao == '':

        opcao = " "

    if opcao != "1" and opcao != "2":

        print("Entrada inválida! Apenas 1 ou 2 são aceitos.")
        tente_novamente("iniciar")



#Esta função é chamada quando o usuário apresenta uma entrada inválida
def tente_novamente(inicio):

    print("------------------------------------------------\n")
    print("Gostaria de tentar novamente?\n")
    print("1 - Sim")
    print("2 - Não\n")

    opcao = input("Digite o número correspondente a sua resposta: ")

    #Aqui estamos dizendo que uma entrada vazia pode ser interpretada como um espaço vazio
    if opcao == '':

        opcao = ' '

    #Aqui estamos verificando se a entrada dada pelo usuário é um número inteiro
    try:
        int(opcao)
        verificacao_numero_inteiro = True

    except ValueError:
        verificacao_numero_inteiro = False

    #Se a entrada for um número inteiro, então...
    if verificacao_numero_inteiro == True:

        #Transformamos a entrada que estaa no formato de escrita em numeral
        opcaoConvert = int(opcao)

        #Se o usuário escolher a opção 1, então o mandamos de volta para o menu inicial
        if opcaoConvert == 1:

            ajustar('iniciar')

        #Se o usuário escolher a opção 2, então encerramos o programa
        if opcaoConvert == 2:

            print("------------------------------------------")
            print("Programa encerrado.")
            print("------------------------------------------")
            exit()

        #Se a entrada não for nem 1 e nem 2, então chamamos novamente a função de erro "tente_novamente"
        if opcaoConvert != 1 and opcaoConvert != 2:

            print("*********************************\n")
            print("Entrada Inválida!")
            print("Você só pode digitar 1 ou 2.")
            print("*********************************\n")

            tente_novamente('iniciar')

    #Se a entrada não for um número inteiro, então chamamos a função de erro 'tente_novamente'
    if verificacao_numero_inteiro == False:

            print("*********************************\n")
            print("Entrada Inválida!")
            print("Você só pode digitar 1 ou 2.\n")
            print("*********************************\n")

            tente_novamente('iniciar')


